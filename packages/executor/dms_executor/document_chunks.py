"""Extract + chunk unstructured uploads for the document index (RAG-01).

No LLM. Dense vectors are deterministic char-trigrams in ``embed_meta`` (hybrid
search with lexical — upgrade to pgvector + model_provider when wired).
"""

from __future__ import annotations

import os
from typing import Any


def extract_text(filename: str, data: bytes) -> str:
    """Pull readable text from UNSTRUCTURED workbook/doc bytes."""
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        parts: list[str] = []
        try:
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        parts.append(" ".join(cells))
        finally:
            wb.close()
        return "\n".join(parts)
    return data.decode("utf-8", errors="replace")


def chunk_text(text: str, *, max_chars: int = 800) -> list[str]:
    """Paragraph-pack chunking — lexical retrieve can refine later (RAG-02)."""
    raw = (text or "").replace("\r\n", "\n").strip()
    if not raw:
        return []
    paragraphs = [p.strip() for p in raw.split("\n\n") if p.strip()]
    if not paragraphs:
        paragraphs = [ln.strip() for ln in raw.split("\n") if ln.strip()]
    out: list[str] = []
    buf = ""
    for para in paragraphs:
        if len(para) > max_chars:
            if buf:
                out.append(buf)
                buf = ""
            for i in range(0, len(para), max_chars):
                out.append(para[i : i + max_chars])
            continue
        candidate = f"{buf}\n\n{para}".strip() if buf else para
        if len(candidate) <= max_chars:
            buf = candidate
        else:
            if buf:
                out.append(buf)
            buf = para
    if buf:
        out.append(buf)
    return out


def index_unstructured_upload(
    *,
    filename: str,
    data: bytes,
    space_id: str,
    blob_key: str | None,
    database_url: str | None = None,
    tenant_id: str | None = None,
) -> dict[str, Any]:
    """Write space-scoped chunks. Returns receipt fields or raises ValueError."""
    conninfo = database_url or os.environ.get("DATABASE_URL")
    if not conninfo:
        raise ValueError("database_url_required")
    if not space_id:
        raise ValueError("space_id_required")
    tid = tenant_id or os.environ.get(
        "DMS_TENANT_ID", "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"
    )
    text = extract_text(filename, data)
    chunks = chunk_text(text)
    if not chunks:
        raise ValueError("no_extractable_text")

    from dms_core.control_plane.document_chunks import index_document

    source_id, count = index_document(
        conninfo,
        tenant_id=tid,
        space_id=space_id,
        ref=filename,
        chunks=chunks,
        blob_key=blob_key,
    )
    return {"source_id": source_id, "chunk_count": count, "document_index": "indexed"}

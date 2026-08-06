"""Sheet triage classifier — header scoring, blank-band, type consistency.

Operates on in-memory grids (list[list[str|None]]). Excel bytes are read via
openpyxl load_workbook(read_only=True) — never create a Workbook, never call
save, never use pandas Excel export writers.

Swap scenario: openpyxl → LibreOffice headless CSV export if binary quirks dominate.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
from collections import Counter
from typing import Any

from dms_core.triage import FileTriageResult, ShapeFingerprint, SheetClass

_NUM_RE = re.compile(r"^[+-]?(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d+)?%?$")
_THOUSANDS_RE = re.compile(r"^\d{1,3}(,\d{3})+(\.\d+)?$")


def _is_blank(cell: Any) -> bool:
    if cell is None:
        return True
    if isinstance(cell, str) and not cell.strip():
        return True
    return False


def _cell_str(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


def _is_numericish(text: str) -> bool:
    if not text:
        return False
    t = text.replace(" ", "")
    return bool(_NUM_RE.match(t))


def _row_nonblank(row: list[Any]) -> list[str]:
    return [_cell_str(c) for c in row if not _is_blank(c)]


def _row_text_ratio(row: list[Any]) -> tuple[float, int, int]:
    cells = _row_nonblank(row)
    if not cells:
        return 0.0, 0, 0
    text_n = sum(1 for c in cells if not _is_numericish(c))
    num_n = len(cells) - text_n
    return text_n / len(cells), text_n, num_n


def _row_uniqueness(row: list[Any]) -> float:
    cells = _row_nonblank(row)
    if not cells:
        return 0.0
    return len(set(cells)) / len(cells)


def score_header_row(grid: list[list[Any]], max_scan: int = 30) -> list[tuple[int, float]]:
    """Return (row_index, score) candidates — higher is more header-like."""
    scores: list[tuple[int, float]] = []
    limit = min(len(grid), max_scan)
    for i in range(limit):
        row = grid[i]
        cells = _row_nonblank(row)
        if len(cells) < 2:
            continue
        text_ratio, text_n, _num_n = _row_text_ratio(row)
        uniq = _row_uniqueness(row)
        # Prefer earlier rows slightly; headers are usually near the top
        position_bonus = max(0.0, 1.0 - (i / max(limit, 1))) * 0.15
        # Header: mostly text labels, high uniqueness
        score = (text_ratio * 0.55) + (uniq * 0.30) + position_bonus
        if text_n >= 2 and text_ratio >= 0.5:
            scores.append((i, score))
    scores.sort(key=lambda x: x[1], reverse=True)
    return scores


def _blank_band_regions(grid: list[list[Any]]) -> list[tuple[int, int]]:
    """Return contiguous non-blank row ranges (start, end exclusive)."""
    regions: list[tuple[int, int]] = []
    start: int | None = None
    for i, row in enumerate(grid):
        blank = all(_is_blank(c) for c in row)
        if not blank and start is None:
            start = i
        elif blank and start is not None:
            regions.append((start, i))
            start = None
    if start is not None:
        regions.append((start, len(grid)))
    return regions


def _type_signature_for_column(values: list[str]) -> str:
    kinds: Counter[str] = Counter()
    for v in values:
        if not v:
            kinds["null"] += 1
        elif _THOUSANDS_RE.match(v.replace(" ", "")) or _is_numericish(v):
            kinds["num"] += 1
        else:
            kinds["text"] += 1
    if not kinds:
        return "empty"
    # dominant non-null
    non_null = {k: n for k, n in kinds.items() if k != "null"}
    if not non_null:
        return "null"
    return max(non_null, key=non_null.get)  # type: ignore[arg-type]


def _column_type_consistency(grid: list[list[Any]], header_idx: int) -> tuple[float, bool]:
    """Return (consistency 0-1, has_thousands_text_numbers)."""
    if header_idx >= len(grid) - 1:
        return 1.0, False
    data = grid[header_idx + 1 :]
    # trim trailing note-like rows (mostly long text, few cols)
    width = max((len(r) for r in grid[header_idx : header_idx + 1]), default=0)
    cols = max(width, max((len(r) for r in data), default=0))
    if cols == 0:
        return 1.0, False
    consistencies: list[float] = []
    has_thousands = False
    for c in range(cols):
        vals = [
            _cell_str(r[c]) if c < len(r) else ""
            for r in data
            if not all(_is_blank(x) for x in r)
        ]
        nonempty = [v for v in vals if v]
        if not nonempty:
            continue
        for v in nonempty:
            if _THOUSANDS_RE.match(v.replace(" ", "")):
                has_thousands = True
        sigs = [_type_signature_for_column([v]) for v in nonempty]
        if not sigs:
            continue
        dominant = Counter(sigs).most_common(1)[0][1]
        consistencies.append(dominant / len(sigs))
    if not consistencies:
        return 1.0, has_thousands
    # Use the worst column — one dirty column makes the sheet dirty
    return min(consistencies), has_thousands


def _looks_unstructured(grid: list[list[Any]]) -> bool:
    if not grid:
        return True
    nonempty_rows = [r for r in grid if _row_nonblank(r)]
    if not nonempty_rows:
        return True
    widths = [len(_row_nonblank(r)) for r in nonempty_rows]
    avg_w = sum(widths) / len(widths)
    cells = sum(widths)
    # Tiny note / single token sheet
    if cells <= 2 and avg_w <= 1.0:
        return True
    long_text = 0
    for r in nonempty_rows:
        for c in _row_nonblank(r):
            if len(c) > 40 and not _is_numericish(c):
                long_text += 1
    if avg_w <= 1.5 and long_text >= max(2, len(nonempty_rows) // 2):
        return True
    if cells <= 3 and len(nonempty_rows) >= 3 and long_text >= 2:
        return True
    return False


def _has_trailing_notes(grid: list[list[Any]], header_idx: int) -> bool:
    if header_idx >= len(grid) - 2:
        return False
    # After a solid block, a short wide-gap then a sparse text row
    data = grid[header_idx + 1 :]
    if len(data) < 3:
        return False
    # Find last dense row vs sparse
    dense = 0
    for r in data:
        if len(_row_nonblank(r)) >= 2:
            dense += 1
        else:
            # single cell long text after dense rows
            cells = _row_nonblank(r)
            if dense >= 2 and cells and len(cells[0]) > 20:
                return True
    return False


def _merged_header_heuristic(grid: list[list[Any]], header_idx: int) -> bool:
    """Title/merged row above header: a single-cell title within 3 rows above."""
    if header_idx <= 0:
        return False
    hdr = _row_nonblank(grid[header_idx])
    if len(hdr) < 2:
        return False
    for back in range(1, min(4, header_idx + 1)):
        above = _row_nonblank(grid[header_idx - back])
        if len(above) == 1 and len(above[0]) > 3:
            return True
    return False


def shape_fingerprint(grid: list[list[Any]], header_idx: int | None) -> ShapeFingerprint:
    if header_idx is None or header_idx >= len(grid):
        col_count = max((len(r) for r in grid), default=0)
        header_text = ""
        type_sig = "unknown"
    else:
        header = [_cell_str(c) for c in grid[header_idx]]
        while header and header[-1] == "":
            header.pop()
        col_count = len(header)
        header_text = "|".join(header).lower()
        consistency, _ = _column_type_consistency(grid, header_idx)
        # per-col signature string
        data = grid[header_idx + 1 :]
        sigs = []
        for c in range(col_count):
            vals = [
                _cell_str(r[c]) if c < len(r) else ""
                for r in data
                if not all(_is_blank(x) for x in r)
            ]
            sigs.append(_type_signature_for_column(vals))
        type_sig = ",".join(sigs) if sigs else f"consistency:{consistency:.2f}"
    h = hashlib.sha256(header_text.encode("utf-8")).hexdigest()[:16]
    return ShapeFingerprint(col_count=col_count, header_text_hash=h, type_signature=type_sig)


_FIXES = {
    "TABULAR_CLEAN": "Ready for bronze ingest as-is.",
    "TABULAR_DIRTY": (
        "Normalize types (strip thousands separators), drop trailing notes, then re-ingest."
    ),
    "MULTI_TABLE": "Split the sheet into one table per blank-band region before ingest.",
    "HEADERLESS": "Add a header row with column names, or map columns in Repair Desk.",
    "UNSTRUCTURED": "Routes to blob tier + document index — will not become a table.",
}


def classify_grid(
    grid: list[list[Any]],
    *,
    file: str = "sheet",
    sheet: str | None = None,
) -> FileTriageResult:
    if not grid or all(all(_is_blank(c) for c in r) for r in grid):
        return FileTriageResult(
            file=file,
            sheet=sheet,
            classification=SheetClass.UNSTRUCTURED,
            reason="empty_or_blank_sheet",
            fix=_FIXES["UNSTRUCTURED"],
            confidence=1.0,
            shape_fingerprint=shape_fingerprint(grid, None),
        )

    if _looks_unstructured(grid):
        return FileTriageResult(
            file=file,
            sheet=sheet,
            classification=SheetClass.UNSTRUCTURED,
            reason="prose_or_notes_not_tabular",
            fix=_FIXES["UNSTRUCTURED"],
            confidence=0.9,
            shape_fingerprint=shape_fingerprint(grid, None),
        )

    regions = _blank_band_regions(grid)
    # MULTI_TABLE: two+ substantial regions
    substantial = [r for r in regions if (r[1] - r[0]) >= 2]
    if len(substantial) >= 2:
        # confirm each region has header-like first row
        multi_ok = True
        for start, end in substantial[:3]:
            sub = grid[start:end]
            if not score_header_row(sub, max_scan=min(5, len(sub))):
                # still multi if blank band separates dense blocks
                if len(_row_nonblank(sub[0])) < 2 and len(sub) > 2:
                    multi_ok = True
        if multi_ok and len(substantial) >= 2:
            hdr_scores = score_header_row(grid)
            header_idx = hdr_scores[0][0] if hdr_scores else substantial[0][0]
            return FileTriageResult(
                file=file,
                sheet=sheet,
                classification=SheetClass.MULTI_TABLE,
                reason=f"blank_band_regions:{len(substantial)}",
                fix=_FIXES["MULTI_TABLE"],
                confidence=0.85,
                header_row=header_idx,
                shape_fingerprint=shape_fingerprint(grid, header_idx),
            )

    hdr_scores = score_header_row(grid)
    if not hdr_scores:
        # numeric/text grid without labels
        return FileTriageResult(
            file=file,
            sheet=sheet,
            classification=SheetClass.HEADERLESS,
            reason="no_header_row_scored",
            fix=_FIXES["HEADERLESS"],
            confidence=0.8,
            shape_fingerprint=shape_fingerprint(grid, None),
        )

    header_idx, hdr_score = hdr_scores[0]
    consistency, has_thousands = _column_type_consistency(grid, header_idx)
    trailing = _has_trailing_notes(grid, header_idx)
    merged = _merged_header_heuristic(grid, header_idx)
    title_above = merged

    dirty_reasons: list[str] = []
    if has_thousands:
        dirty_reasons.append("numbers_as_text_thousands_separators")
    if trailing:
        dirty_reasons.append("trailing_notes_rows")
    if title_above:
        dirty_reasons.append("title_or_merged_row_above_header")
    if consistency < 0.9:
        dirty_reasons.append("column_type_inconsistency")

    fp = shape_fingerprint(grid, header_idx)
    if dirty_reasons:
        return FileTriageResult(
            file=file,
            sheet=sheet,
            classification=SheetClass.TABULAR_DIRTY,
            reason="+".join(dirty_reasons),
            fix=_FIXES["TABULAR_DIRTY"],
            confidence=min(0.95, 0.55 + hdr_score * 0.3),
            header_row=header_idx,
            shape_fingerprint=fp,
        )

    return FileTriageResult(
        file=file,
        sheet=sheet,
        classification=SheetClass.TABULAR_CLEAN,
        reason="header_and_types_ok",
        fix=_FIXES["TABULAR_CLEAN"],
        confidence=min(0.99, 0.7 + hdr_score * 0.25),
        header_row=header_idx,
        shape_fingerprint=fp,
    )


def parse_csv_grid(data: bytes, *, encoding: str = "utf-8") -> list[list[Any]]:
    if data.startswith(b"\xef\xbb\xbf"):
        data = data[3:]
    text = data.decode(encoding, errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def parse_xlsx_grids(data: bytes) -> list[tuple[str, list[list[Any]]]]:
    """Read-only openpyxl — never creates/saves a workbook."""
    from openpyxl import load_workbook

    bio = io.BytesIO(data)
    wb = load_workbook(bio, read_only=True, data_only=True)
    out: list[tuple[str, list[list[Any]]]] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            grid: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                grid.append(list(row))
            out.append((name, grid))
    finally:
        wb.close()
    return out


def classify_bytes(
    *,
    filename: str,
    data: bytes,
) -> list[FileTriageResult]:
    lower = filename.lower()
    if lower.endswith(".csv") or lower.endswith(".tsv"):
        grid = parse_csv_grid(data)
        return [classify_grid(grid, file=filename, sheet=None)]
    if lower.endswith((".xlsx", ".xlsm")):
        try:
            sheets = parse_xlsx_grids(data)
        except Exception as exc:  # noqa: BLE001
            return [
                FileTriageResult(
                    file=filename,
                    sheet=None,
                    classification=SheetClass.UNSTRUCTURED,
                    reason=f"xlsx_read_error:{exc}"[:180],
                    fix="Re-export sheet as CSV or fix the workbook, then re-ingest.",
                    confidence=0.5,
                )
            ]
        if not sheets:
            return [
                FileTriageResult(
                    file=filename,
                    sheet=None,
                    classification=SheetClass.UNSTRUCTURED,
                    reason="xlsx_no_sheets",
                    fix=_FIXES["UNSTRUCTURED"],
                    confidence=1.0,
                )
            ]
        return [classify_grid(g, file=filename, sheet=name) for name, g in sheets]
    if lower.endswith(".xls"):
        return [
            FileTriageResult(
                file=filename,
                sheet=None,
                classification=SheetClass.UNSTRUCTURED,
                reason="xls_legacy_unsupported",
                fix="Save as .xlsx or CSV, then re-ingest.",
                confidence=1.0,
            )
        ]
    # plaintext / unknown
    try:
        text = data.decode("utf-8", errors="replace")
        if "\t" in text or "," in text.splitlines()[0] if text.splitlines() else False:
            grid = parse_csv_grid(data)
            return [classify_grid(grid, file=filename)]
        lines = [[ln] for ln in text.splitlines()]
        return [classify_grid(lines, file=filename)]
    except Exception:  # noqa: BLE001
        return [
            FileTriageResult(
                file=filename,
                sheet=None,
                classification=SheetClass.UNSTRUCTURED,
                reason="unreadable_bytes",
                fix=_FIXES["UNSTRUCTURED"],
                confidence=0.7,
            )
        ]

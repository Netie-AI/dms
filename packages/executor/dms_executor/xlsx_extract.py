"""Extract a Copilot-built workbook into the Space artifact store (dms#31).

Byte-faithful copy of a workbook produced elsewhere. Not authoring one
(hard rule 5). This module copies bytes and only load_workbook(read_only).

Kind is ``xlsx_result`` (resulting Copilot artifact). That is not
AirGPT #20 ingested-originals / ``data_sources.kind='document'``.
"""

from __future__ import annotations

import hashlib
import json
import os
import uuid
from io import BytesIO
from pathlib import Path
from typing import Any

from dms_executor.batch_ingest import _blob_put
from dms_executor.demo_warehouse import warehouse_path

# FRTR families on the STORED artifact. Missing any => incomplete, not green.
REQUIRED_FAMILIES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("cover", ("cover",)),
    ("ontime_export", ("ontime export", "on-time export", "on time export")),
    ("analysis", ("analysis",)),
    ("presentation_chart", ("presentation chart",)),
)

KIND = "xlsx_result"


def _norm_sheet(name: str) -> str:
    return " ".join(
        (name or "").lower().replace("-", " ").replace("_", " ").split()
    )


def _family_present(normed_names: list[str], needles: tuple[str, ...]) -> bool:
    for raw in normed_names:
        padded = f" {raw} "
        for needle in needles:
            if f" {needle} " in padded:
                return True
    return False


def sheet_families(sheet_names: list[str]) -> dict[str, Any]:
    """Report which FRTR families the stored workbook has. Never silently green."""
    normed = [_norm_sheet(n) for n in sheet_names]
    missing: list[str] = []
    present: list[str] = []
    for family, needles in REQUIRED_FAMILIES:
        if _family_present(normed, needles):
            present.append(family)
        else:
            missing.append(family)
    return {
        "sheets": list(sheet_names),
        "present_families": present,
        "missing_families": missing,
        "complete": not missing,
    }


def prove_stored_sheets(data: bytes) -> dict[str, Any]:
    """Sheet-presence proof on STORED bytes (not the input path)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - named fail, never skip
        raise AssertionError(f"openpyxl unavailable for sheet proof: {exc}") from exc

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()
    return sheet_families(names)


def _artifact_dir(root: Path) -> Path:
    d = root / "artifacts"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _write_sidecar(root: Path, record: dict[str, Any]) -> Path:
    dest = _artifact_dir(root) / f"{record['id']}.json"
    dest.write_text(json.dumps(record, indent=2), encoding="utf-8")
    return dest


def get_artifact(artifact_id: str, *, root: Path | None = None) -> dict[str, Any] | None:
    """Later reveal: load the durable sidecar row by id."""
    blob_root = root if root is not None else warehouse_path().parent
    path = _artifact_dir(blob_root) / f"{artifact_id}.json"
    if not path.is_file():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def extract_resulting_xlsx(
    workbook_path: str | Path,
    *,
    space_id: str,
    root: Path | None = None,
    tenant_id: str | None = None,
    database_url: str | None = None,
) -> dict[str, Any]:
    """Copy a resulting workbook into the artifact store. Return durable path/id.

    Does not drive Copilot, Pointer, Excel, or MCP. Caller supplies the path
    XLSX-ORCH-10 would produce. Live Copilot input still depends on dms#30.
    """
    if not space_id:
        raise ValueError("space_id_required")
    src = Path(workbook_path)
    if not src.is_file():
        raise FileNotFoundError(f"workbook_not_found: {src}")

    data = src.read_bytes()
    digest = hashlib.sha256(data).hexdigest()
    blob_root = root if root is not None else warehouse_path().parent
    stored_path = _blob_put(digest, data, root=blob_root)
    stored = Path(stored_path)
    stored_bytes = stored.read_bytes()
    stored_digest = hashlib.sha256(stored_bytes).hexdigest()
    if stored_digest != digest:
        raise RuntimeError(
            f"store_truncated: input sha256={digest} stored sha256={stored_digest}"
        )

    families = prove_stored_sheets(stored_bytes)
    artifact_id = str(uuid.uuid4())
    tid = tenant_id or os.environ.get(
        "DMS_TENANT_ID", "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"
    )
    record: dict[str, Any] = {
        "id": artifact_id,
        "path": stored_path,
        "space_id": space_id,
        "tenant_id": tid,
        "kind": KIND,
        "sha256": digest,
        "origin_path": str(src),
        "sheets": families["sheets"],
        "present_families": families["present_families"],
        "missing_families": families["missing_families"],
        "complete": families["complete"],
        "store": "sidecar",
    }
    _write_sidecar(blob_root, record)

    conninfo = database_url if database_url is not None else os.environ.get("DATABASE_URL")
    if conninfo:
        try:
            from dms_core.control_plane.space_artifacts import register_artifact

            register_artifact(
                conninfo,
                tenant_id=tid,
                space_id=space_id,
                artifact_id=artifact_id,
                blob_key=stored_path,
                sha256=digest,
                origin_path=str(src),
                sheets=families["sheets"],
                complete=families["complete"],
                missing_families=families["missing_families"],
            )
            record["store"] = "postgres+sidecar"
            _write_sidecar(blob_root, record)
        except Exception as exc:  # noqa: BLE001 — named, never silent green
            record["store"] = "sidecar_only"
            record["postgres_error"] = f"{type(exc).__name__}: {exc}"[:180]
            _write_sidecar(blob_root, record)

    return record

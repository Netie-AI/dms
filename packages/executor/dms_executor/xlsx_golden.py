"""FRTR golden OnTime avg for the Copilot path (dms#32 / XLSX-ORCH-12).

Asserted on numbers visible in Analysis / OnTime Export of a Pointer->Excel
Copilot workbook. MCP-primary and openpyxl-as-primary are refused as the
Demo-2 golden source.

Tolerance is documented here until the founder pins exact rounding.
This module never authors a workbook (hard rule 5).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

# FRTR 00027 OnTime=true ballpark (ticket dms#32).
GOLDEN_AVG = 300.27
GOLDEN_ONTIME = 184005
GOLDEN_TOTAL = 200000
# Founder-agreed rounding has not been pinned tighter than the ticket's "~".
AVG_TOLERANCE = 0.05
COUNT_TOLERANCE = 50

ALLOWED_ORIGINS = frozenset({"copilot_pointer"})
REFUSED_ORIGINS = frozenset({"mcp", "mcp_primary", "user-excel", "openpyxl", "openpyxl_as_primary"})


class GoldenRejected(AssertionError):
    """Named golden miss. Never skip, never xfail."""


def _norm(text: str) -> str:
    return " ".join((text or "").lower().replace("-", " ").replace("_", " ").split())


def _is_analysis_or_export(name: str) -> bool:
    n = _norm(name)
    return "analysis" in n or "ontime export" in n or "on time export" in n


def _closer(value: float, target: float, tol: float) -> bool:
    return abs(value - target) <= tol


def verify_frtr_golden(
    workbook_path: str | Path,
    *,
    origin: str,
) -> dict[str, Any]:
    """Read Analysis/Export numbers. Fail loudly on wrong origin, filter, or column."""
    origin_key = _norm(str(origin or "")).replace(" ", "_")
    if origin_key in REFUSED_ORIGINS or origin_key not in ALLOWED_ORIGINS:
        raise GoldenRejected(
            f"golden_refused_origin: {origin!r} "
            "(Demo-2 golden is Copilot path / Pointer paste only; "
            "MCP-primary and openpyxl-as-primary are not the source)"
        )

    src = Path(workbook_path)
    if not src.is_file():
        raise GoldenRejected(f"workbook_not_found: {src}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - named fail, never skip
        raise GoldenRejected(f"openpyxl unavailable for golden proof: {exc}") from exc

    wb = load_workbook(src, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        numbers: list[float] = []
        scanned = 0
        for name in sheet_names:
            if not _is_analysis_or_export(name):
                continue
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, bool):
                        continue
                    if isinstance(cell, (int, float)):
                        numbers.append(float(cell))
                        scanned += 1
    finally:
        wb.close()

    if scanned == 0:
        raise GoldenRejected(
            "golden_no_numbers: Analysis/Export had no numeric cells "
            "(wrong filter, formula not cached, or Copilot path did not land)"
        )

    avg_hit = next((n for n in numbers if _closer(n, GOLDEN_AVG, AVG_TOLERANCE)), None)
    ontime_hit = next(
        (n for n in numbers if _closer(n, float(GOLDEN_ONTIME), float(COUNT_TOLERANCE))),
        None,
    )
    total_hit = next(
        (n for n in numbers if _closer(n, float(GOLDEN_TOTAL), float(COUNT_TOLERANCE))),
        None,
    )
    misses: list[str] = []
    if avg_hit is None:
        misses.append(
            f"avg_cost not ~{GOLDEN_AVG} (+/- {AVG_TOLERANCE}); saw {numbers[:12]}"
        )
    if ontime_hit is None:
        misses.append(
            f"ontime_count not ~{GOLDEN_ONTIME} (+/- {COUNT_TOLERANCE})"
        )
    if total_hit is None:
        misses.append(
            f"total_count not ~{GOLDEN_TOTAL} (+/- {COUNT_TOLERANCE})"
        )
    if misses:
        raise GoldenRejected("golden_mismatch: " + "; ".join(misses))

    return {
        "ok": True,
        "origin": "copilot_pointer",
        "path": str(src),
        "avg_cost": avg_hit,
        "ontime_count": int(ontime_hit) if ontime_hit is not None else None,
        "total_count": int(total_hit) if total_hit is not None else None,
        "tolerance": {
            "avg": AVG_TOLERANCE,
            "count": COUNT_TOLERANCE,
            "note": "ticket ballpark until founder pins exact rounding",
        },
        "sheets": sheet_names,
    }


__all__ = [
    "ALLOWED_ORIGINS",
    "AVG_TOLERANCE",
    "COUNT_TOLERANCE",
    "GOLDEN_AVG",
    "GOLDEN_ONTIME",
    "GOLDEN_TOTAL",
    "GoldenRejected",
    "verify_frtr_golden",
]

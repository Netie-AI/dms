"""Governed cross-check of an AirGPT Copilot prompt pack (dms#30).

Consumes the AirGPT XLSX-ORCH-01 pack. Strengthens without becoming a second
orchestrator. Emits a Pointer-paste payload. Does not paste, does not drive
Excel Copilot, and does not call MCP as primary.

Status stays ``awaiting_pointer_receipt`` until Pointer posts a resulting
workbook path. Extract is dms#31. Golden numbers are dms#32.
"""

from __future__ import annotations

import json
import re
import time
from pathlib import Path
from typing import Any

from dms_executor.demo_warehouse import warehouse_path

HONESTY = (
    "DMS governed cross-check only. Pointer owns Excel Copilot paste. "
    "DMS does not run Copilot or MCP paste."
)
HANDOFF_HONESTY = (
    "No Pointer paste in this repo. Status stays awaiting_pointer_receipt "
    "until Pointer posts a resulting workbook path. Extract is dms#31."
)

REQUIRED_FAMILIES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("cover", ("cover",)),
    ("ontime_export", ("ontime export", "on-time export", "on time export")),
    ("analysis", ("analysis",)),
    ("presentation_chart", ("presentation chart",)),
)

class PackRejected(ValueError):
    """Named refusal. Never a silent green on a weak or mis-owned pack."""


def _orch_dir(root: Path | None = None) -> Path:
    blob_root = root if root is not None else warehouse_path().parent
    dest = blob_root / "xlsx_orch"
    dest.mkdir(parents=True, exist_ok=True)
    return dest


def _norm(text: str) -> str:
    return " ".join((text or "").lower().replace("-", " ").replace("_", " ").split())


def _blob(parts: list[Any]) -> str:
    return " ".join(_norm(str(p)) for p in parts if p is not None)


def _family_present(normed_names: list[str], needles: tuple[str, ...]) -> bool:
    for raw in normed_names:
        padded = f" {raw} "
        for needle in needles:
            if f" {needle} " in padded:
                return True
    return False


def sheet_families(sheet_names: list[str]) -> dict[str, Any]:
    normed = [_norm(n) for n in sheet_names]
    missing: list[str] = []
    present: list[str] = []
    for family, needles in REQUIRED_FAMILIES:
        if _family_present(normed, needles):
            present.append(family)
        else:
            missing.append(family)
    return {
        "sheets": list(sheet_names),
        "present_families": present,
        "missing_families": missing,
        "complete": not missing,
    }


def unwrap_pack(payload: dict[str, Any]) -> dict[str, Any]:
    """Accept AirGPT identify_and_pack response or the inner pack dict."""
    if not isinstance(payload, dict):
        raise PackRejected("missing_pack: payload is not an object")
    inner = payload.get("pack")
    if isinstance(inner, dict) and (
        inner.get("steps") is not None or inner.get("expected_result_sheets") is not None
    ):
        pack = dict(inner)
        if payload.get("pack_id") and not pack.get("pack_id"):
            pack["airgpt_pack_id"] = payload["pack_id"]
        if payload.get("workbook") and not pack.get("workbook"):
            pack["workbook"] = payload["workbook"]
        return pack
    return dict(payload)


def inspect_result_workbook(path: str | Path) -> dict[str, Any]:
    """Sheet-presence smoke on a Pointer-posted path. Read-only. No authoring."""
    p = Path(path)
    if not p.is_file():
        return {
            "ok": False,
            "error": f"workbook_not_found: {p}",
            "path": str(p),
            "complete": False,
        }
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - named fail, never skip
        raise AssertionError(f"openpyxl unavailable for sheet proof: {exc}") from exc

    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()
    families = sheet_families(names)
    return {
        "ok": families["complete"],
        "path": str(p.resolve()),
        "error": None if families["complete"] else (
            "missing sheet families: " + ",".join(families["missing_families"])
        ),
        **families,
    }


def _require_pointer_owner(pack: dict[str, Any]) -> None:
    owner = _norm(str(pack.get("paste_owner") or ""))
    if owner in {"mcp", "user excel", "user-excel", "excel mcp"}:
        raise PackRejected(
            "mcp_as_primary: paste_owner is MCP; Demo-2 paste owner is Pointer"
        )
    if owner and owner != "pointer":
        raise PackRejected(
            f"paste_owner_not_pointer: paste_owner={pack.get('paste_owner')!r}"
        )
    if not owner:
        raise PackRejected("paste_owner_not_pointer: paste_owner missing")


def _require_columns(pack: dict[str, Any]) -> dict[str, Any]:
    wb = pack.get("workbook") if isinstance(pack.get("workbook"), dict) else {}
    ontime = str(wb.get("ontime_col") or "").strip()
    cost = str(wb.get("cost_col") or "").strip()
    if not ontime:
        raise PackRejected("missing_ontime_col: pack is too weak to cross-check")
    if not cost:
        raise PackRejected("missing_cost_col: pack is too weak to cross-check")
    return wb


def _require_steps(pack: dict[str, Any]) -> list[dict[str, Any]]:
    steps = pack.get("steps")
    if not isinstance(steps, list) or len(steps) < 4:
        raise PackRejected("steps_incomplete: need Cover / Export / Analysis / Chart")
    blob = _blob(
        [
            s.get("sheet")
            for s in steps
            if isinstance(s, dict)
        ]
        + [
            s.get("intent")
            for s in steps
            if isinstance(s, dict)
        ]
        + [
            s.get("formula_hint")
            for s in steps
            if isinstance(s, dict)
        ]
    )
    if "cover" not in blob:
        raise PackRejected("steps_missing_cover")
    if "ontime" not in blob and "on time" not in blob:
        raise PackRejected("steps_missing_ontime_filter")
    if "averageif" not in blob and "filter" not in blob:
        raise PackRejected("steps_missing_averageif_or_filter")
    if "export" not in blob:
        raise PackRejected("steps_missing_export")
    if "chart" not in blob and "ppt" not in blob:
        raise PackRejected("steps_missing_chart")
    return [s for s in steps if isinstance(s, dict)]


def _require_expected_sheets(pack: dict[str, Any]) -> list[str]:
    raw = pack.get("expected_result_sheets") or []
    if not isinstance(raw, list) or not raw:
        raise PackRejected("missing_expected_sheets")
    names = [str(x) for x in raw]
    families = sheet_families(names)
    if not families["complete"]:
        raise PackRejected(
            "missing_expected_sheets: "
            + ",".join(families["missing_families"])
        )
    return names


def _workbook_path(pack: dict[str, Any], override: str | None) -> str:
    wb = pack.get("workbook") if isinstance(pack.get("workbook"), dict) else {}
    cand = (override or "").strip() or str(wb.get("path") or "").strip()
    return cand


def strengthen_pack(pack: dict[str, Any]) -> dict[str, Any]:
    """Overlay DMS governance. Do not rewrite the AirGPT steps from scratch."""
    wb = pack.get("workbook") if isinstance(pack.get("workbook"), dict) else {}
    ontime = str(wb.get("ontime_col") or "OnTime")
    cost = str(wb.get("cost_col") or "Cost")
    src = str(wb.get("source_sheet") or "the fact sheet")
    overlay = [
        "DMS governed overlay (do not invent a second orchestrator):",
        f"Filter {ontime}=TRUE on {src} only. Do not average a Summary sheet.",
        f"On Analysis, AVERAGEIF or AVERAGE of FILTER for {cost} on that set.",
        "State OnTime count vs total row count in cells Pointer can later extract.",
        "Do not invent rows, ranks, or KPIs that are not in the source sheet.",
        "Do not treat MCP as the paste path. Pointer pastes into Excel Copilot.",
        "Cover is provenance (ask, source file, filter, output sheet names).",
    ]
    out = dict(pack)
    out["dms_governance"] = overlay
    out["paste_owner"] = "pointer"
    out["cross_check_owner"] = "dms"
    out["dms_role"] = "governed_cross_check"
    out["not_doing"] = sorted(
        set(list(pack.get("not_doing") or []) + [
            "pointer_paste",
            "excel_copilot_drive",
            "mcp_user_excel_primary",
            "openpyxl_as_primary",
        ])
    )
    out["honesty"] = HONESTY
    return out


def build_paste_text(pack: dict[str, Any]) -> str:
    """Text Pointer pastes into Excel Copilot. DMS does not paste it."""
    lines = list(pack.get("dms_governance") or [])
    lines.append("")
    for step in pack.get("steps") or []:
        if not isinstance(step, dict):
            continue
        n = step.get("n") or "?"
        sheet = step.get("sheet") or ""
        lines.append(f"{n}. [{sheet}] {step.get('intent') or ''}")
        hint = step.get("formula_hint")
        if hint:
            lines.append(f"   Formula hint: {hint}")
    return "\n".join(lines).strip()


def _new_pack_id(airgpt_id: str | None) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", str(airgpt_id or "pack"))[:40]
    return f"dms_orch_{int(time.time() * 1000)}_{slug.strip('_') or 'pack'}"


def save_receipt(rec: dict[str, Any], *, root: Path | None = None) -> dict[str, Any]:
    dest = _orch_dir(root) / f"{rec['pack_id']}.json"
    dest.write_text(json.dumps(rec, indent=2), encoding="utf-8")
    rec["receipt_path"] = str(dest)
    return rec


def load_receipt(pack_id: str, *, root: Path | None = None) -> dict[str, Any] | None:
    fp = _orch_dir(root) / f"{str(pack_id or '').strip()}.json"
    if not fp.is_file():
        return None
    return json.loads(fp.read_text(encoding="utf-8"))


def crosscheck_airgpt_pack(
    payload: dict[str, Any],
    *,
    workbook_path: str | None = None,
    root: Path | None = None,
) -> dict[str, Any]:
    """Validate + strengthen an AirGPT pack. Hand off to Pointer. Do not paste."""
    pack = unwrap_pack(payload)
    if not pack:
        raise PackRejected("missing_pack")
    _require_pointer_owner(pack)
    if str(pack.get("airgpt_role") or "") not in {"candidate_pack_only", ""}:
        # Empty is tolerated on a trimmed fixture; a steal is not.
        if str(pack.get("airgpt_role") or "").lower() in {
            "orchestrator",
            "pointer",
            "copilot_driver",
        }:
            raise PackRejected("airgpt_role_steals_excel_drive")
    wb = _require_columns(pack)
    steps = _require_steps(pack)
    expected = _require_expected_sheets(pack)
    path = _workbook_path(pack, workbook_path)
    if not path:
        raise PackRejected("workbook_path_missing")
    src = Path(path)
    if not src.is_file():
        raise PackRejected(f"workbook_not_found: {src}")

    strengthened = strengthen_pack(pack)
    strengthened["steps"] = steps
    strengthened["expected_result_sheets"] = expected
    if isinstance(strengthened.get("workbook"), dict):
        strengthened["workbook"] = {**wb, "path": str(src)}
    paste_text = build_paste_text(strengthened)
    pack_id = _new_pack_id(str(pack.get("airgpt_pack_id") or pack.get("pack_id") or ""))
    rec = {
        "ok": True,
        "pack_id": pack_id,
        "airgpt_pack_id": pack.get("airgpt_pack_id") or payload.get("pack_id"),
        "status": "awaiting_pointer_receipt",
        "paste_owner": "pointer",
        "workbook_path": str(src),
        "paste_text": paste_text,
        "expected_result_sheets": expected,
        "pack": strengthened,
        "result_path": None,
        "result_sheets": None,
        "not_doing": strengthened["not_doing"],
        "honesty": HANDOFF_HONESTY,
        "created_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
    }
    return save_receipt(rec, root=root)


def receive_pointer_result(
    pack_id: str,
    result_path: str,
    *,
    root: Path | None = None,
) -> dict[str, Any]:
    """Accept a Pointer-posted resulting workbook path. Do not extract (dms#31)."""
    rec = load_receipt(pack_id, root=root)
    if rec is None:
        raise PackRejected(f"pack_not_found: {pack_id}")
    inspected = inspect_result_workbook(result_path)
    rec["result_path"] = inspected.get("path")
    rec["result_sheets"] = inspected
    rec["status"] = (
        "pointer_received" if inspected.get("ok") else "pointer_received_incomplete_sheets"
    )
    rec["honesty"] = HANDOFF_HONESTY
    return save_receipt(rec, root=root)


__all__ = [
    "HANDOFF_HONESTY",
    "HONESTY",
    "PackRejected",
    "build_paste_text",
    "crosscheck_airgpt_pack",
    "inspect_result_workbook",
    "load_receipt",
    "receive_pointer_result",
    "sheet_families",
    "strengthen_pack",
    "unwrap_pack",
]

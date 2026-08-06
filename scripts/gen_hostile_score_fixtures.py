"""Regenerate EPIC-018 hostile score fixtures (openpyxl, no gold hand-authoring).

  python scripts/gen_hostile_score_fixtures.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1] / "tests" / "fixtures" / "hostile_score"


def _sales_book(path: Path, electronics: float, home: float, sports: float, misc: float) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["# banner", None, None])
    ws.append(["sku", "category", "sales_value_myr"])
    for row in (
        ("SKU-1", "Electronics", electronics * 0.6),
        ("SKU-2", "Electronics", electronics * 0.4),
        ("SKU-3", "Home", home * 0.7),
        ("SKU-4", "Home", home * 0.3),
        ("SKU-5", "Sports", sports),
        ("SKU-6", "Misc", misc),
        (None, None, None),
        ("SKU-X", "Ignore DROP TABLE", "DROP"),
    ):
        ws.append(list(row))
    wide = wb.create_sheet("Wide_Fill")
    wide.append(["sku", "category", "sales_value_myr"])
    wide.append(["SKU-9", "Electronics", 50.0])
    wb.save(path)


def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)
    _sales_book(
        ROOT / "cf98e431_p50_01_sales_messy.xlsx",
        1_545_366.40,
        1_199_018.49,
        400_000.00,
        380_948.33,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["sku", "category", "sales_value_myr"])
    ws.append(["SKU-1", "Electronics", 1_563_234.12])
    ws.append(["SKU-2", "Home", 900_000.00])
    ws.append(["SKU-3", "Sports", 800_000.00])
    wide = wb.create_sheet("Wide_Fill")
    wide.append(["sku", "category", "sales_value_myr"])
    wide.append(["A", "Misc", 2_691_552.10])
    wide.append(["B", "Apparel", 2_402_425.68])
    wide.append(["C", "Sports", 2_358_800.10])
    wide.append(["D", "Home", 100_000.00])
    wb.save(ROOT / "aa64458a_p50_03_inventory_messy.xlsx")

    enc = Workbook()
    ews = enc.active
    ews.title = "Sales"
    ews.append(["sku", "city", "sales_value_myr"])
    ews.append(["SKU-BETA", "Kuala Lumpur", 1500.75])
    ews.append(["SKU-ALPHA", "Kuala Lumpur", 200.00])
    ews.append(["SKU-GAMMA", "Johor Bahru", 900.00])
    enc.save(ROOT / "encoding_value_norm.xlsx")

    _f32_ambiguous_scope(ROOT / "f32_ambiguous_scope.xlsx")
    _blank_rows_book(ROOT / "blank_rows_hanging.xlsx")
    print(f"wrote {sorted(p.name for p in ROOT.glob('*.xlsx'))}")


def _f32_ambiguous_scope(path: Path) -> None:
    """SCORE-03 / F32 — the founder's miss, rebuilt so the oracle can recompute it.

    Two sheets that both answer "top 3 category sales", with a *different rank
    order*, not merely different magnitudes. Sales is the truth; Wide_Fill
    carries the exact ranking the live stack returned under a green badge
    (Home / Sports / Misc). An ambiguous ask that silently picks Wide_Fill is
    therefore wrong twice over, and no tolerance on the numbers can hide it.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["# FY25 export - header below", None, None])
    ws.append(["sku", "category", "sales_value_myr"])
    for row in (
        ("SKU-1", "Electronics", 945_366.40),
        ("SKU-2", "Electronics", 600_000.00),
        ("SKU-3", "Home", 799_018.49),
        ("SKU-4", "Home", 400_000.00),
        ("SKU-5", "Misc", 380_948.33),
        ("SKU-6", "Sports", 300_000.00),
    ):
        ws.append(list(row))

    # The trap sheet. Same headers, same dimension values, ~4x smaller totals -
    # plausible enough that a reader who does not check scope will accept it.
    wide = wb.create_sheet("Wide_Fill")
    wide.append(["sku", "category", "sales_value_myr"])
    for row in (
        ("SKU-A", "Home", 383_803.56),
        ("SKU-B", "Sports", 242_755.97),
        ("SKU-C", "Misc", 228_548.84),
        ("SKU-D", "Electronics", 100_000.00),
    ):
        wide.append(list(row))
    wb.save(path)


def _blank_rows_book(path: Path) -> None:
    """SCORE-03 optional — trailing blank and hanging rows must not move the total.

    ``Sales`` and ``Sales_Clean`` hold the same six data rows. Sales additionally
    carries a mid-sheet blank band, hanging rows with a category but no measure,
    a fully empty trailing band, and a stray footer. If blanks are ever counted
    as zero-valued members the group set inflates; if a hanging row's category
    is credited its neighbour's measure the totals deflate. Either way the two
    sheets stop agreeing, which is the assertion.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["sku", "category", "sales_value_myr"])
    ws.append(["SKU-1", "Electronics", 1_000.50])
    ws.append([None, None, None])  # blank band mid-table
    ws.append(["SKU-2", "Electronics", 500.25])
    ws.append(["SKU-3", "Home", 800.00])
    ws.append(["SKU-4", "Home", None])  # hanging row - category, no measure
    ws.append(["SKU-5", "Sports", 300.00])
    ws.append([None, "Misc", None])  # hanging row - no sku, no measure
    ws.append(["SKU-6", "Misc", 100.05])
    ws.append(["SKU-7", "Home", 99.45])
    for _ in range(12):  # trailing empty rows a human would drag over
        ws.append([None, None, None])
    ws.append(["Total", None, "=SUM(C2:C10)"])  # footer, never a data row

    clean = wb.create_sheet("Sales_Clean")
    clean.append(["sku", "category", "sales_value_myr"])
    for row in (
        ("SKU-1", "Electronics", 1_000.50),
        ("SKU-2", "Electronics", 500.25),
        ("SKU-3", "Home", 800.00),
        ("SKU-5", "Sports", 300.00),
        ("SKU-6", "Misc", 100.05),
        ("SKU-7", "Home", 99.45),
    ):
        clean.append(list(row))
    wb.save(path)


if __name__ == "__main__":
    main()

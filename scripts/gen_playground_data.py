"""Generate playground sample workbooks + a short policy note for L3 RAG.

  python scripts/gen_playground_data.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1] / "playground" / "data"


def _sales(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["sku", "category", "city", "sales_value_myr", "units"])
    rows = [
        ("SKU-ALPHA", "Electronics", "Kuala Lumpur", 420_500.00, 120),
        ("SKU-BETA", "Electronics", "Kuala Lumpur", 380_250.50, 95),
        ("SKU-GAMMA", "Home", "Johor Bahru", 290_100.00, 200),
        ("SKU-DELTA", "Home", "Shah Alam", 210_000.00, 150),
        ("SKU-EPSILON", "Sports", "Penang", 175_800.25, 80),
        ("SKU-ZETA", "Misc", "Klang", 95_000.00, 40),
        ("SKU-ETA", "Apparel", "Kuala Lumpur", 310_400.00, 110),
        ("SKU-THETA", "Apparel", "Johor Bahru", 140_000.00, 70),
    ]
    for r in rows:
        ws.append(list(r))
    wide = wb.create_sheet("Wide_Fill")
    wide.append(["sku", "category", "sales_value_myr"])
    # Trap sheet: different ranking on purpose (F26 class)
    wide.append(["W1", "Misc", 2_100_000.00])
    wide.append(["W2", "Apparel", 1_900_000.00])
    wide.append(["W3", "Sports", 1_700_000.00])
    wide.append(["W4", "Electronics", 50_000.00])
    wb.save(path)


def _ops(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"
    ws.append(["shipment_id", "sku", "status", "city", "delay_days"])
    for row in (
        ("SH-01", "SKU-BETA", "delayed", "Kuala Lumpur", 3),
        ("SH-02", "SKU-ALPHA", "delivered", "Kuala Lumpur", 0),
        ("SH-03", "SKU-GAMMA", "delayed", "Johor Bahru", 5),
        ("SH-04", "SKU-ETA", "in_transit", "Penang", 1),
        ("SH-05", "SKU-DELTA", "delivered", "Shah Alam", 0),
    ):
        ws.append(list(row))
    wb.save(path)


def _policy_note(path: Path) -> None:
    path.write_text(
        """Playground policy note (unstructured — L3 RAG only)

Late delivery penalty: 5,000.00 MYR per late shipment after the grace window.
Grace window: 2 days.

Do not treat this note as a warehouse total. Category sales live in the Sales sheet.
Contact: ops@example.local

Synonym hint for stewards (not SQL): "product family" means category.
""",
        encoding="utf-8",
    )


def _returns_policy(path: Path) -> None:
    path.write_text(
        """# Playground returns policy (unstructured - L3 only)

Late delivery penalty: 5,000.00 MYR per late shipment (Clause 7.2).

Do not invent category sales totals from this document.
Numeric aggregates must come from the Sales workbook via SQL (L2) or a
certified/trusted asset (L0/L1). This file is for citation / prose only.
""",
        encoding="utf-8",
    )


def main() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)
    _sales(ROOT / "pg_sales.xlsx")
    _ops(ROOT / "pg_shipments.xlsx")
    _policy_note(ROOT / "pg_policy_note.txt")
    _returns_policy(ROOT / "pg_returns_policy.md")
    print(f"wrote {sorted(p.name for p in ROOT.iterdir())}")


if __name__ == "__main__":
    main()

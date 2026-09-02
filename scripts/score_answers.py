"""EPIC-018 - coverage / precision instrument for the DMS answer path.

A single "accuracy" score is retired (F26). It is the metric that let a
top-scoring RAG build hand a client three wrong category totals under a green
badge. Two numbers replace it:

  precision-on-answered   of confidently-badged answers, the share matching the
                          oracle on BOTH magnitude and rank order.
                          Target 100.00 pct. This is a law, not a goal.
  coverage                share of the pack answered rather than abstained.
                          Grows every wave. Never bought with precision.

The oracle is recomputed from the source workbooks on every run, with openpyxl
- never read from a stored gold file (R-0009: never hand-author a generated
artifact), and never computed by the DuckDB path it is grading (R-0003: the
adversary is not the verifier).

  Oracle only, no stack needed:
    python scripts/score_answers.py --docs <dir> --oracle-only

  Score a live stack:
    python scripts/score_answers.py --docs <dir> --space <space_id>

Exit 1 on a single confidently-wrong answer, regardless of coverage.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Money tolerance: a cent, or a relative whisker for large sums.
ABS_TOL = 0.011
REL_TOL = 1e-6

# Header can sit below an export-dump banner, so scan a few rows for it.
HEADER_SCAN_ROWS = 8


def _close(a: float, b: float) -> bool:
    return abs(a - b) <= ABS_TOL or abs(a - b) / max(abs(b), 1e-9) <= REL_TOL


def _as_number(cell: Any) -> float | None:
    if isinstance(cell, bool) or cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    text = str(cell).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def grouped_totals(path: Path, sheet: str, dim: str, measure: str) -> dict[str, float]:
    """Sum ``measure`` by ``dim`` over one sheet of one workbook.

    Deliberately narrow: one file, one sheet. A total that silently spans
    sheets or workbooks is the defect this instrument exists to catch, so the
    oracle cannot express one.

    Rows whose measure is not numeric are dropped - that is what quarantines
    the blank band and the injection trailer the fixtures carry, and it matches
    what Excel's SUM does when a human checks the number by hand.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise KeyError(f"{path.name} has no sheet {sheet!r}; has {wb.sheetnames}")
        ws = wb[sheet]

        header: dict[str, int] | None = None
        totals: dict[str, float] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if header is None:
                if i >= HEADER_SCAN_ROWS:
                    break
                labels = {
                    str(c).strip().lower(): idx
                    for idx, c in enumerate(row)
                    if c is not None and str(c).strip()
                }
                if dim.lower() in labels and measure.lower() in labels:
                    header = {dim: labels[dim.lower()], measure: labels[measure.lower()]}
                continue

            d_idx, m_idx = header[dim], header[measure]
            if d_idx >= len(row) or m_idx >= len(row):
                continue
            key = row[d_idx]
            val = _as_number(row[m_idx])
            if val is None or key is None or not str(key).strip():
                continue
            totals[str(key).strip()] = totals.get(str(key).strip(), 0.0) + val

        if header is None:
            raise KeyError(f"{path.name}::{sheet} has no {dim!r}+{measure!r} header row")
        return totals
    finally:
        wb.close()


def oracle_top_n(path: Path, sheet: str, dim: str, measure: str, n: int) -> list[tuple[str, float]]:
    ranked = sorted(grouped_totals(path, sheet, dim, measure).items(), key=lambda kv: -kv[1])
    return ranked[:n]


def oracle_eq_filter(
    path: Path, sheet: str, filter_col: str, filter_value: str, measure: str
) -> list[tuple[str, float]]:
    """Exact-match filter total. Encoding mismatch is the defect under test."""
    totals = grouped_totals(path, sheet, filter_col, measure)
    if filter_value not in totals:
        raise KeyError(
            f"{path.name}::{sheet} has no {filter_col}={filter_value!r}; "
            f"keys={sorted(totals)}"
        )
    return [(filter_value, totals[filter_value])]


def resolve_oracle(case: dict[str, Any], path: Path) -> list[tuple[str, float]]:
    kind = str(case.get("oracle") or "top_n")
    if kind == "eq_filter":
        return oracle_eq_filter(
            path,
            str(case["sheet"]),
            str(case["filter_col"]),
            str(case["filter_value"]),
            str(case["measure"]),
        )
    return oracle_top_n(
        path,
        str(case["sheet"]),
        str(case["dim"]),
        str(case["measure"]),
        int(case["top_n"]),
    )


# --------------------------------------------------------------------------
# Question pack. Scope is declared, because a question with no declared scope
# has no single right answer - Wide_Fill totals run ~2x the Sales totals in the
# same workbook, so "which sheet" moves the answer by 100 pct.
#
# fail_mode_today documents how a weak stack fails green TODAY. Pack staying
# red on live score is intended until value-norm / trusted assets land (F31).
# Default docs dir for local break tests: tests/fixtures/hostile_score/
# --------------------------------------------------------------------------
QUESTION_PACK: list[dict[str, Any]] = [
    {
        "id": "sales01_cat_top3",
        "workbook": "cf98e431_p50_01_sales_messy.xlsx",
        "sheet": "Sales",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": "cross-file merge trap twin - silent union with inventory03",
        "question": (
            "In cf98e431_p50_01_sales_messy.xlsx sheet Sales, what are the top 3 "
            "categories by sales_value_myr?"
        ),
    },
    {
        "id": "sales01_widefill_top3",
        "workbook": "cf98e431_p50_01_sales_messy.xlsx",
        "sheet": "Wide_Fill",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": "wrong-sheet scope - answers Sales while Wide_Fill is named",
        "question": (
            "In cf98e431_p50_01_sales_messy.xlsx sheet Wide_Fill, what are the top 3 "
            "categories by sales_value_myr?"
        ),
    },
    {
        # Adversarial pair with the case above: same question shape, different
        # workbook. A system that merges sources returns one number for both.
        "id": "inventory03_cat_top3",
        "workbook": "aa64458a_p50_03_inventory_messy.xlsx",
        "sheet": "Sales",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": "cross-file merge trap - one ranking for two workbooks",
        "question": (
            "In aa64458a_p50_03_inventory_messy.xlsx sheet Sales, what are the top 3 "
            "categories by sales_value_myr?"
        ),
    },
    {
        # Same workbook, different sheet. This is the pair the client's demo
        # actually got wrong: it answered from Wide_Fill while he read Sales.
        "id": "inventory03_widefill_top3",
        "workbook": "aa64458a_p50_03_inventory_messy.xlsx",
        "sheet": "Wide_Fill",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": "wrong-sheet scope - answers Sales while Wide_Fill is named",
        "question": (
            "In aa64458a_p50_03_inventory_messy.xlsx sheet Wide_Fill, what are the top 3 "
            "categories by sales_value_myr?"
        ),
    },
    {
        "id": "synonym_product_family_top3",
        "workbook": "cf98e431_p50_01_sales_messy.xlsx",
        "sheet": "Sales",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": (
            "synonym/acronym - 'product family' / 'cat' not wired to category; "
            "abstain or wrong filter (not regex-fixed)"
        ),
        "question": (
            "In cf98e431_p50_01_sales_messy.xlsx sheet Sales, top 3 product families "
            "by MYR sales (cat / product line synonym for category)?"
        ),
    },
    {
        "id": "wrong_sheet_named_sales_trap",
        "workbook": "aa64458a_p50_03_inventory_messy.xlsx",
        "sheet": "Sales",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": (
            "wrong-sheet scope - question says Sales but mentions Wide_Fill numbers "
            "context; confident answer from Wide_Fill ranking"
        ),
        "question": (
            "Using aa64458a_p50_03_inventory_messy.xlsx, on the Sales sheet only "
            "(ignore Wide_Fill), what are the top 3 categories by sales_value_myr?"
        ),
    },
    {
        "id": "rag_sum_invite_no_sql",
        "workbook": "aa64458a_p50_03_inventory_messy.xlsx",
        "sheet": "Wide_Fill",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": (
            "RAG-sum trap - F26 class: invent partial totals from retrieved prose; "
            "correct E9 behaviour is abstain (coverage cost), never confident WRONG"
        ),
        "question": (
            "From the documents about aa64458a_p50_03_inventory_messy.xlsx Wide_Fill, "
            "summarize the top 3 category sales totals without running warehouse SQL."
        ),
    },
    {
        "id": "empty_filter_beta_vs_sku_beta",
        "workbook": "encoding_value_norm.xlsx",
        "sheet": "Sales",
        "oracle": "eq_filter",
        "filter_col": "sku",
        "filter_value": "SKU-BETA",
        "measure": "sales_value_myr",
        "fail_mode_today": (
            "empty-filter green - literal filter BETA matches nothing; truth is "
            "SKU-BETA (hard rule 12); zero/empty under confident badge is P0"
        ),
        "question": (
            "In encoding_value_norm.xlsx sheet Sales, what is total sales_value_myr "
            "for sku BETA?"
        ),
    },
    {
        "id": "empty_filter_kl_vs_kuala_lumpur",
        "workbook": "encoding_value_norm.xlsx",
        "sheet": "Sales",
        "oracle": "eq_filter",
        "filter_col": "city",
        "filter_value": "Kuala Lumpur",
        "measure": "sales_value_myr",
        "fail_mode_today": (
            "empty-filter green - city 'KL' vs stored 'Kuala Lumpur'; "
            "confident empty is false precision"
        ),
        "question": (
            "In encoding_value_norm.xlsx sheet Sales, total sales_value_myr for city KL?"
        ),
    },
    {
        "id": "exact_sku_beta_total",
        "workbook": "encoding_value_norm.xlsx",
        "sheet": "Sales",
        "oracle": "eq_filter",
        "filter_col": "sku",
        "filter_value": "SKU-BETA",
        "measure": "sales_value_myr",
        "fail_mode_today": (
            "exact stored encoding still PathNotAllowed FROM transactions; "
            "must certify SKU-BETA from the named sheet, never rewrite BETA"
        ),
        "question": (
            "In encoding_value_norm.xlsx sheet Sales, what is total sales_value_myr "
            "for sku SKU-BETA?"
        ),
    },
    {
        "id": "exact_city_kuala_lumpur_total",
        "workbook": "encoding_value_norm.xlsx",
        "sheet": "Sales",
        "oracle": "eq_filter",
        "filter_col": "city",
        "filter_value": "Kuala Lumpur",
        "measure": "sales_value_myr",
        "fail_mode_today": (
            "exact city encoding still PathNotAllowed FROM transactions; "
            "must certify Kuala Lumpur from the named sheet, never rewrite KL"
        ),
        "question": (
            "In encoding_value_norm.xlsx sheet Sales, total sales_value_myr "
            "for city Kuala Lumpur?"
        ),
    },
    {
        # F32 - the founder's miss. Every other case in this pack names its
        # sheet; this one deliberately does not, because that is the whole
        # defect: nothing in the question says which sheet, so the stack picks
        # one and states the result with a confident badge. The workbook is
        # named only so the oracle has a file to recompute from - the sheet, the
        # dimension column, and the measure all stay unsaid, and "categoty" is
        # left misspelled the way it was typed.
        #
        # The oracle is declared on Sales even though the question does not say
        # Sales. That is not the instrument taking a side in an ambiguity: an
        # answer with no defensible scope is not correct-but-unlucky, it is
        # unfounded, and the only honest response to it is to abstain. Abstain
        # costs coverage here and nothing else. What must never pass is the
        # Wide_Fill ranking under a green badge.
        "id": "f32_ambiguous_categoty_top3",
        "workbook": "f32_ambiguous_scope.xlsx",
        "sheet": "Sales",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": (
            "F32 underspecified scope + typo - no sheet named, 'categoty' misspelled; "
            "confident Wide_Fill ranking (Home 383,803.56 / Sports 242,755.97 / "
            "Misc 228,548.84) against Sales truth (Electronics 1,545,366.40 / "
            "Home 1,199,018.49 / Misc 380,948.33) - wrong rank AND wrong magnitude "
            "under green; abstain is the correct answer until VQ-01 / E9-02 land"
        ),
        "question": "In f32_ambiguous_scope.xlsx, show top 3 categoty sales",
    },
    {
        # Optional case allowed by SCORE-03, under EPIC-018 only. Not a scope
        # trap - a parsing trap. The sheet carries a mid-table blank band,
        # hanging rows with a category but no measure, twelve trailing empty
        # rows, and a formula footer. Sales_Clean holds the same six data rows
        # with none of that, so the two top-3s must be identical.
        "id": "blank_hanging_rows_top3",
        "workbook": "blank_rows_hanging.xlsx",
        "sheet": "Sales",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": (
            "blank/hanging rows treated as data - trailing empties counted as zero "
            "members inflate the group set, or a hanging row is credited its "
            "neighbour's measure and deflates the total; either way the messy sheet "
            "stops agreeing with Sales_Clean while the badge stays green"
        ),
        "question": (
            "In blank_rows_hanging.xlsx sheet Sales, what are the top 3 categories "
            "by sales_value_myr?"
        ),
    },
    {
        "id": "malay_paraphrase_top3",
        "workbook": "cf98e431_p50_01_sales_messy.xlsx",
        "sheet": "Sales",
        "dim": "category",
        "measure": "sales_value_myr",
        "top_n": 3,
        "fail_mode_today": (
            "multilingual/Malay paraphrase - miss route or invent helpful totals; "
            "must abstain or match Sales oracle, never fabricate"
        ),
        "question": (
            "Dalam fail cf98e431_p50_01_sales_messy.xlsx helaian Sales, apakah 3 "
            "kategori teratas mengikut sales_value_myr?"
        ),
    },
]


def extract_ranking(env: dict[str, Any]) -> list[tuple[str, float]] | None:
    """Read the ranking the envelope actually claims, from executed rows.

    E9 guarantees a non-abstained numeric answer carries executed rows, so rows
    are the honest place to read the claim from. Row order is the ranking - a
    top-N answer whose ORDER BY is wrong is wrong even when every total is right.
    """
    rows = env.get("rows") or []
    if not rows:
        return None
    first = rows[0]
    dim_key = next((k for k, v in first.items() if isinstance(v, str)), None)
    num_key = next(
        (k for k, v in first.items() if isinstance(v, (int, float)) and not isinstance(v, bool)),
        None,
    )
    if dim_key is None or num_key is None:
        return None
    out: list[tuple[str, float]] = []
    for r in rows:
        key, val = r.get(dim_key), r.get(num_key)
        if isinstance(key, str) and isinstance(val, (int, float)) and not isinstance(val, bool):
            out.append((key.strip(), float(val)))
    return out or None


def judge(env: dict[str, Any], expected: list[tuple[str, float]]) -> tuple[str, str]:
    """Classify one answer. Returns (outcome, detail).

    outcome is one of: abstained | correct | WRONG
    Only 'correct' and 'WRONG' touch precision. 'abstained' costs coverage
    only - refusing is always allowed, being confidently wrong never is.
    """
    if env.get("abstained"):
        return "abstained", str(env.get("text") or "")[:120]

    claimed = extract_ranking(env)
    if claimed is None:
        return "WRONG", "confident badge with no executed rows to read a ranking from"

    claimed = claimed[: len(expected)]
    if len(claimed) < len(expected):
        return "WRONG", f"returned {len(claimed)} ranked rows, expected {len(expected)}"

    exp_keys = [k.lower() for k, _ in expected]
    got_keys = [k.lower() for k, _ in claimed]
    if exp_keys != got_keys:
        return "WRONG", f"rank order {got_keys} != oracle {exp_keys}"

    for (_, got), (key, want) in zip(claimed, expected):
        if not _close(got, want):
            return "WRONG", f"{key}: answered {got:,.2f}, oracle {want:,.2f}"

    return "correct", ", ".join(f"{k}={v:,.2f}" for k, v in claimed)


def ask_live(
    question: str,
    space_id: str,
    timeout: float,
    *,
    grounded_tables: list[str] | None = None,
) -> dict[str, Any]:
    import httpx

    base = os.environ.get("DMS_URL", "http://127.0.0.1:8090").rstrip("/")
    payload: dict[str, Any] = {"question": question, "space_id": space_id}
    if grounded_tables:
        payload["grounded_tables"] = grounded_tables
    resp = httpx.post(
        f"{base}/v1/chat/ask",
        json=payload,
        timeout=timeout,
    )
    if resp.status_code == 403:
        detail = resp.json().get("detail") if resp.headers.get("content-type", "").startswith("application/json") else {}
        if not isinstance(detail, dict):
            detail = {"message": str(detail)}
        code = str(detail.get("code") or "")
        if code in {"grounding_not_grantable", "outside_grounded_scope"}:
            return {
                "abstained": True,
                "badge": "ABSTAIN",
                "text": str(detail.get("message") or code),
                "rows": [],
                "values": [],
            }
    resp.raise_for_status()
    return dict(resp.json())


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Metrics\n"
            "  precision-on-answered  share of confident answers matching oracle "
            "(magnitude + rank). Target 100.00 pct.\n"
            "  coverage               share answered rather than abstained. "
            "Never buys a WRONG.\n"
            "\n"
            "Exit codes\n"
            "  0  PASS — 0 confidently wrong (abstains allowed)\n"
            "  1  FAIL — at least one confident WRONG, or ask/stack error\n"
            "\n"
            "Examples\n"
            "  python scripts/score_answers.py "
            "--docs tests/fixtures/hostile_score --oracle-only\n"
            "  set DMS_URL=http://127.0.0.1:8090\n"
            "  python scripts/score_answers.py "
            "--docs tests/fixtures/hostile_score --space <space_id>\n"
            "\n"
            "Oracle is openpyxl recomputed each run (never a hand gold file, "
            "never DuckDB). Hostile pack may stay red until value-norm / "
            "EPIC-019 — that is intended.\n"
        ),
    )
    ap.add_argument(
        "--docs",
        required=True,
        type=Path,
        help="directory holding the workbooks (e.g. tests/fixtures/hostile_score)",
    )
    ap.add_argument(
        "--space",
        help="Space id to ask against via POST /v1/chat/ask; omit with --oracle-only",
    )
    ap.add_argument(
        "--oracle-only",
        action="store_true",
        help="print the recomputed openpyxl oracle + fail_mode_today and stop",
    )
    ap.add_argument("--timeout", type=float, default=180.0)
    args = ap.parse_args(argv)

    if not args.oracle_only and not args.space:
        ap.error("--space is required unless --oracle-only is given")

    cases: list[tuple[dict[str, Any], list[tuple[str, float]]]] = []
    for case in QUESTION_PACK:
        path = args.docs / str(case["workbook"])
        if not path.is_file():
            print(f"FAIL missing workbook: {path}")
            return 1
        cases.append((case, resolve_oracle(case, path)))

    print("=== ORACLE (recomputed this run, openpyxl, one file + one sheet each) ===")
    for case, expected in cases:
        scope = f"{case['workbook']}::{case['sheet']}"
        body = "  ".join(f"{k}={v:,.2f}" for k, v in expected)
        fail = str(case.get("fail_mode_today") or "")
        print(f"  {case['id']:<36} {scope}")
        print(f"  {'':<36} {body}")
        if fail:
            print(f"  {'':<36} fail_mode_today: {fail}")

    # Merge trap: these two scopes MUST disagree. Shared paraphrases of the same
    # scope (synonym / Malay) intentionally share an oracle and are not a WARN.
    by_id = {case["id"]: exp for case, exp in cases}

    def _fingerprint(pairs: list[tuple[str, float]]) -> tuple[Any, ...]:
        return tuple(k.lower() for k, _ in pairs) + tuple(round(v, 2) for _, v in pairs)

    a, b = by_id.get("sales01_cat_top3"), by_id.get("inventory03_cat_top3")
    if a is not None and b is not None:
        fa, fb = _fingerprint(a), _fingerprint(b)
        print(f"\n  merge-trap oracles distinct: {fa != fb}")
        if fa == fb:
            print("  WARN sales01 and inventory03 share an answer - merge trap is blind")

    # F32 scope trap: the ambiguity only bites if the two candidate sheets
    # disagree. If they ever converge, a stack could pick the wrong sheet and
    # still score correct - the trap would be reporting green while blind.
    f32 = by_id.get("f32_ambiguous_categoty_top3")
    if f32 is not None:
        wide = oracle_top_n(
            args.docs / "f32_ambiguous_scope.xlsx", "Wide_Fill", "category",
            "sales_value_myr", 3,
        )
        distinct = _fingerprint(f32) != _fingerprint(wide)
        print(f"  F32 scope-trap oracles distinct: {distinct}")
        print(f"    Sales     {'  '.join(f'{k}={v:,.2f}' for k, v in f32)}")
        print(f"    Wide_Fill {'  '.join(f'{k}={v:,.2f}' for k, v in wide)}")
        if not distinct:
            print("  WARN F32 Sales and Wide_Fill agree - the scope trap is blind")
            return 1

    # Blank/hanging rows must be invisible to the oracle. Asserted here rather
    # than only in pytest so a regenerated fixture cannot quietly break it.
    blanks = by_id.get("blank_hanging_rows_top3")
    if blanks is not None:
        clean = oracle_top_n(
            args.docs / "blank_rows_hanging.xlsx", "Sales_Clean", "category",
            "sales_value_myr", 3,
        )
        agree = _fingerprint(blanks) == _fingerprint(clean)
        print(f"  blank/hanging rows do not move the oracle: {agree}")
        if not agree:
            print(f"  FAIL messy {_fingerprint(blanks)} != clean {_fingerprint(clean)}")
            return 1

    if args.oracle_only:
        return 0

    print(f"\n=== SCORING (space={args.space}) ===")
    answered = correct = wrong = 0
    failures: list[str] = []
    for case, expected in cases:
        try:
            from dms_executor.bronze import bronze_table_for_sheet

            grounded = [
                bronze_table_for_sheet(str(case["workbook"]), str(case.get("sheet") or "Sales"))
            ]
            env = ask_live(
                str(case["question"]),
                str(args.space),
                args.timeout,
                grounded_tables=grounded,
            )
        except Exception as exc:  # noqa: BLE001 - a dead stack is a run failure, not a score
            print(f"  {case['id']:<32} ERROR  {type(exc).__name__}: {exc}")
            failures.append(f"{case['id']}: ask failed ({type(exc).__name__})")
            continue

        outcome, detail = judge(env, expected)
        badge = env.get("badge")
        print(f"  {case['id']:<32} {outcome:<10} [{badge}] {detail}")
        if outcome == "abstained":
            continue
        answered += 1
        if outcome == "correct":
            correct += 1
        else:
            wrong += 1
            failures.append(f"{case['id']}: {detail}")

    total = len(cases)
    precision = (correct / answered * 100.0) if answered else 100.0
    coverage = answered / total * 100.0 if total else 0.0

    print("\n=== RESULT ===")
    print(f"  precision-on-answered  {precision:6.2f} pct   ({correct}/{answered})")
    print(f"  coverage               {coverage:6.2f} pct   ({answered}/{total})")

    art = Path(os.environ.get("DMS_SCORE_DIR") or (Path(__file__).resolve().parents[1] / ".tmp"))
    art.mkdir(parents=True, exist_ok=True)
    (art / "score_hostile.json").write_text(
        json.dumps(
            {
                "kind": "dms.score_answers",
                "pack": "hostile_score",
                "precision_on_answered": round(precision, 2),
                "coverage_pct": round(coverage, 2),
                "correct": correct,
                "answered": answered,
                "wrong": wrong,
                "total": total,
                "abstained": total - answered,
                "passed": wrong == 0 and not failures,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    if wrong:
        print(f"\nFAIL {wrong} confidently wrong answer(s). Coverage does not buy this back.")
        for f in failures:
            print(f"  - {f}")
        return 1
    if failures:
        print("\nFAIL run incomplete:")
        for f in failures:
            print(f"  - {f}")
        return 1

    print("\nPASS 0 confidently wrong.")
    if coverage < 100.0:
        print(f"     {total - answered} abstained - raise coverage by curation, not by loosening.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

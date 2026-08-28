"""XLSX-ORCH-12 / dms#32 -- FRTR golden OnTime avg on the Copilot path.

The checker must be able to fail. Fixture workbooks here prove the gate, they
are not the Demo-2 golden source (openpyxl-as-primary is refused).
"""

from __future__ import annotations

from pathlib import Path

import pytest

try:
    from openpyxl import Workbook
except ImportError as exc:  # named fail, never skip
    raise AssertionError(f"openpyxl required for xlsx golden gate: {exc}") from exc

from dms_executor.xlsx_golden import (
    GOLDEN_AVG,
    GOLDEN_ONTIME,
    GOLDEN_TOTAL,
    GoldenRejected,
    verify_frtr_golden,
)


def _analysis_wb(path: Path, *, avg: float, ontime: int, total: int) -> Path:
    wb = Workbook()
    cover = wb.active
    assert cover is not None
    cover.title = "Cover"
    export = wb.create_sheet("OnTime Export")
    export.append(["OnTime", "Cost"])
    analysis = wb.create_sheet("Analysis")
    analysis.append(["avg_cost", avg])
    analysis.append(["ontime_count", ontime])
    analysis.append(["total_count", total])
    wb.create_sheet("Presentation Chart")
    wb.save(path)
    wb.close()
    return path


def test_golden_matches_frtr_ballpark_on_copilot_origin(tmp_path: Path) -> None:
    src = _analysis_wb(
        tmp_path / "copilot_result.xlsx",
        avg=GOLDEN_AVG,
        ontime=GOLDEN_ONTIME,
        total=GOLDEN_TOTAL,
    )
    out = verify_frtr_golden(src, origin="copilot_pointer")
    assert out["ok"] is True
    assert abs(out["avg_cost"] - GOLDEN_AVG) <= 0.05
    assert out["ontime_count"] == GOLDEN_ONTIME
    assert out["total_count"] == GOLDEN_TOTAL


def test_wrong_avg_fails_loudly(tmp_path: Path) -> None:
    src = _analysis_wb(
        tmp_path / "wrong_filter.xlsx",
        avg=199.99,
        ontime=GOLDEN_ONTIME,
        total=GOLDEN_TOTAL,
    )
    with pytest.raises(GoldenRejected, match="avg_cost"):
        verify_frtr_golden(src, origin="copilot_pointer")


def test_mcp_origin_is_not_the_demo2_golden_source(tmp_path: Path) -> None:
    src = _analysis_wb(
        tmp_path / "mcp.xlsx",
        avg=GOLDEN_AVG,
        ontime=GOLDEN_ONTIME,
        total=GOLDEN_TOTAL,
    )
    with pytest.raises(GoldenRejected, match="golden_refused_origin"):
        verify_frtr_golden(src, origin="mcp")


def test_openpyxl_as_primary_is_not_the_demo2_golden_source(tmp_path: Path) -> None:
    src = _analysis_wb(
        tmp_path / "authored.xlsx",
        avg=GOLDEN_AVG,
        ontime=GOLDEN_ONTIME,
        total=GOLDEN_TOTAL,
    )
    with pytest.raises(GoldenRejected, match="golden_refused_origin"):
        verify_frtr_golden(src, origin="openpyxl_as_primary")


def test_missing_workbook_fails(tmp_path: Path) -> None:
    with pytest.raises(GoldenRejected, match="workbook_not_found"):
        verify_frtr_golden(tmp_path / "nope.xlsx", origin="copilot_pointer")

"""XLSX-ORCH-10 / dms#30 -- consume AirGPT pack, governed cross-check, Pointer handoff.

Does not paste into Pointer, drive Excel Copilot, or treat MCP as primary.
Does not require a live Copilot session. Fixture xlsx is tests-only (hard rule 5).
"""

from __future__ import annotations

from pathlib import Path

import pytest

try:
    from openpyxl import Workbook
except ImportError as exc:  # named fail, never skip
    raise AssertionError(f"openpyxl required for xlsx orch gate: {exc}") from exc

from dms_executor.xlsx_orch import (
    PackRejected,
    crosscheck_airgpt_pack,
    load_receipt,
    receive_pointer_result,
)


F15_ASK = (
    "OnTime=true average cost, export a separate xlsx, and a chart for PPT "
    "on frtr_00027 supply-chain regional"
)


def _source_xlsx(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Carriers"
    ws.append(["Carrier", "Region", "OnTime", "Cost"])
    ws.append(["Acme", "APAC", True, 310.0])
    wb.save(path)
    wb.close()
    return path


def _result_xlsx(path: Path, sheets: list[str]) -> Path:
    wb = Workbook()
    first = wb.active
    assert first is not None
    first.title = sheets[0]
    for name in sheets[1:]:
        wb.create_sheet(name)
    wb.save(path)
    wb.close()
    return path


def _airgpt_pack(src: Path, **overrides: object) -> dict:
    pack = {
        "ok": True,
        "ask": F15_ASK,
        "workbook": {
            "title": "frtr_00027_supply-chain-regional.xlsx",
            "path": str(src),
            "source_id": 9,
            "source_sheet": "Carriers",
            "ontime_col": "OnTime",
            "cost_col": "Cost",
            "region_col": "Region",
            "n_rows": 200000,
        },
        "steps": [
            {
                "n": 1,
                "sheet": "Cover",
                "intent": "Create a Cover sheet that names the ask and the source.",
                "formula_hint": "No aggregate on Cover.",
            },
            {
                "n": 2,
                "sheet": "OnTime Export",
                "intent": "On OnTime Export, spill every Carriers row where OnTime is TRUE.",
                "formula_hint": "=FILTER('Carriers'!A:Z,'Carriers'!OnTime=TRUE)",
            },
            {
                "n": 3,
                "sheet": "Analysis",
                "intent": "On Analysis, compute the average Cost over the OnTime=TRUE set only.",
                "formula_hint": "=AVERAGEIF('Carriers'!OnTime,TRUE,'Carriers'!Cost)",
            },
            {
                "n": 4,
                "sheet": "Presentation Chart",
                "intent": "On Presentation Chart, add a chart of the OnTime analysis for PPT.",
                "formula_hint": "Column chart from Analysis.",
            },
        ],
        "expected_result_sheets": [
            "Cover",
            "OnTime Export",
            "Analysis",
            "Presentation Chart",
        ],
        "paste_owner": "pointer",
        "cross_check_owner": "dms",
        "airgpt_role": "candidate_pack_only",
        "not_doing": ["pointer_paste", "excel_copilot_drive", "mcp_user_excel_primary"],
    }
    pack.update(overrides)
    return pack


def test_crosscheck_accepts_airgpt_pack_and_waits_for_pointer(tmp_path: Path) -> None:
    src = _source_xlsx(tmp_path / "frtr_00027_supply-chain-regional.xlsx")
    out = crosscheck_airgpt_pack(_airgpt_pack(src), root=tmp_path)

    assert out["ok"] is True
    assert out["status"] == "awaiting_pointer_receipt"
    assert out["paste_owner"] == "pointer"
    assert out["result_path"] is None
    assert "Cover" in out["paste_text"]
    assert "AVERAGEIF" in out["paste_text"]
    assert "Pointer owns" in out["honesty"] or "awaiting_pointer_receipt" in out["honesty"]
    assert "pointer_paste" in out["not_doing"]
    assert "mcp_user_excel_primary" in out["not_doing"]
    assert Path(out["workbook_path"]).is_file()
    later = load_receipt(out["pack_id"], root=tmp_path)
    assert later is not None
    assert later["status"] == "awaiting_pointer_receipt"


def test_unwraps_airgpt_identify_response(tmp_path: Path) -> None:
    src = _source_xlsx(tmp_path / "frtr.xlsx")
    payload = {
        "ok": True,
        "status": "identified",
        "pack_id": "orch_airgpt_1",
        "pack": _airgpt_pack(src),
    }
    out = crosscheck_airgpt_pack(payload, root=tmp_path)
    assert out["airgpt_pack_id"] == "orch_airgpt_1"
    assert out["status"] == "awaiting_pointer_receipt"


def test_mcp_as_primary_is_refused(tmp_path: Path) -> None:
    src = _source_xlsx(tmp_path / "frtr.xlsx")
    with pytest.raises(PackRejected, match="mcp_as_primary"):
        crosscheck_airgpt_pack(
            _airgpt_pack(src, paste_owner="mcp"),
            root=tmp_path,
        )


def test_weak_pack_without_ontime_col_is_refused(tmp_path: Path) -> None:
    src = _source_xlsx(tmp_path / "frtr.xlsx")
    pack = _airgpt_pack(src)
    pack["workbook"]["ontime_col"] = ""
    with pytest.raises(PackRejected, match="missing_ontime_col"):
        crosscheck_airgpt_pack(pack, root=tmp_path)


def test_missing_workbook_file_is_refused(tmp_path: Path) -> None:
    ghost = tmp_path / "nope.xlsx"
    with pytest.raises(PackRejected, match="workbook_not_found"):
        crosscheck_airgpt_pack(_airgpt_pack(ghost), root=tmp_path)


def test_pointer_receipt_smokes_sheets_and_does_not_extract(tmp_path: Path) -> None:
    src = _source_xlsx(tmp_path / "frtr.xlsx")
    rec = crosscheck_airgpt_pack(_airgpt_pack(src), root=tmp_path)
    result = _result_xlsx(
        tmp_path / "copilot_result.xlsx",
        ["Cover", "OnTime Export", "OnTime Analysis", "Presentation Chart"],
    )
    posted = receive_pointer_result(rec["pack_id"], str(result), root=tmp_path)
    assert posted["status"] == "pointer_received"
    assert posted["result_path"]
    assert Path(posted["result_path"]).is_file()
    assert posted["result_sheets"]["complete"] is True
    # Extract is dms#31 -- this ticket only persists the path.
    assert "artifact" not in posted
    assert posted.get("kind") != "xlsx_result"


def test_incomplete_pointer_workbook_is_not_greened(tmp_path: Path) -> None:
    src = _source_xlsx(tmp_path / "frtr.xlsx")
    rec = crosscheck_airgpt_pack(_airgpt_pack(src), root=tmp_path)
    thin = _result_xlsx(tmp_path / "thin.xlsx", ["Cover"])
    posted = receive_pointer_result(rec["pack_id"], str(thin), root=tmp_path)
    assert posted["status"] == "pointer_received_incomplete_sheets"
    assert posted["result_sheets"]["complete"] is False
    assert "ontime_export" in posted["result_sheets"]["missing_families"]
